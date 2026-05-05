"""Build KayFit roadmap.xlsx — single flat sheet, user decides priority."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/user/Desktop/КУРСОР/mobileKayfit")
OUT = ROOT / "specs" / "roadmap" / "KayFit_Roadmap.xlsx"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

DONE_FILL = PatternFill("solid", fgColor="D1FAE5")
CODE_ONLY_FILL = PatternFill("solid", fgColor="FEF3C7")
TODO_FILL = PatternFill("solid", fgColor="FEE2E2")
PARTIAL_FILL = PatternFill("solid", fgColor="DBEAFE")


def fill_status(cell, status: str) -> None:
    s = status.lower().strip()
    if s == "done":
        cell.fill = DONE_FILL
    elif s == "code-only":
        cell.fill = CODE_ONLY_FILL
    elif s == "partial":
        cell.fill = PARTIAL_FILL
    elif s in ("open", "not started"):
        cell.fill = TODO_FILL


# ─── Единый список ────────────────────────────────────────────────────────
# Поля: ID, Категория, Задача, Статус, Файлы / источник, Что нужно сделать
#
# Статусы (только эти):
#   Done       — закрыто, ручная регрессия пройдена
#   Code-only  — код написан + юнит-тесты, ручной тест НЕ пройден
#   Partial    — часть закрыта, часть открыта
#   Open       — не реализовано
#   Not started— только сформулировано

ITEMS = [
    # ─ Hotfix P0 — все Code-only до ручной регрессии ─
    ("1.1", "Hotfix", "Сохранение онбординга/сессии", "Code-only",
     "secure_token_storage.dart, onboarding_screen.dart",
     "Manual: Apple Sign-In + kill + relaunch через 5+ мин → Dashboard без логина"),
    ("1.2", "Hotfix", "Постоянный разлогин", "Code-only",
     "auth_provider.dart, api_client.dart",
     "Manual: истёкший access token → действие → silent refresh без логина"),
    ("1.5", "Hotfix", "Зависание AI Data Processing", "Done",
     "ai_consent_screen.dart, ai_consent_provider.dart",
     "Manual PASS: Accept проходит, нет зависания. Таймаут 5→20с + 401-fallback на logout работают"),
    ("2.1", "Hotfix", "Слетающий язык RU↔EN", "Code-only",
     "app.dart, locale_provider.dart, settings_screen.dart, locale_interceptor.dart",
     "Manual: чистая установка → EN; RU через Settings; перезапуск сохраняет выбор"),
    ("3.1", "Hotfix", "Пропадает кнопка «Далее» в онбординге", "Done",
     "onboarding_screen.dart, onboarding_scaffold.dart, ob_gradient_button.dart",
     "Manual PASS на симуляторе iPhone 16 (1.7 + 1.13 + 1.15)"),
    ("4.3", "Hotfix", "Вода 0 ккал", "Code-only",
     "zero_calorie_whitelist.py (бэк), add_meal_sheet.dart",
     "Manual: вода фото → 0; «вода» текстом → 0; «sparkling water» → 0"),
    ("5.1", "Hotfix", "Распознавание не-еды как еды", "Open",
     "ai_service_v2.py, schemas_v2.py, add_meal_sheet.dart",
     "FAIL manual: фото случайной картинки → 'Something went wrong' вместо not-food snackbar. Бэк не возвращает is_food:false ИЛИ валит запрос. Логи в работе"),
    ("7.1", "Hotfix", "Не закрывается клавиатура", "Partial",
     "keyboard_dismisser.dart + 5 экранов",
     "Manual PASS на онбординге (1.6 1.7). AddMeal/EditMeal/Chat/RecognitionResult — не проверены"),
    ("7.2", "Hotfix", "Не уходят bottom sheets", "Code-only",
     "dismissible_sheet_wrapper.dart + 8 шторок",
     "Manual: тап-вне / свайп / X — каждый из 3 способов на каждой шторке"),
    ("7.3", "Hotfix", "Не выйти из AI-чата", "Code-only",
     "chat_screen.dart",
     "Manual: AI-чат → стрелка назад → Dashboard; tab BottomNav из чата"),

    # ─ HIGH warnings из QA — мелкие фиксы по hotfix ─
    ("HIGH-1", "Hotfix follow-up", "ai_consent: лоадер мгновенно вместо 300 мс по ТЗ", "Open",
     "ai_consent_screen.dart:378",
     "Добавить Timer(300ms) перед setConsent ИЛИ согласовать как принятое отступление"),
    ("HIGH-2", "Hotfix follow-up", "ai_consent: Localizations.localeOf вместо localeProvider", "Open",
     "ai_consent_screen.dart:40,446",
     "Заменить 2 вхождения на ref.watch(localeProvider).languageCode == 'ru'"),
    ("HIGH-3", "Hotfix follow-up", "add_meal: нет фронт zero-calorie guard для текстового ввода", "Open",
     "add_meal_sheet.dart, ingredient_v2.dart",
     "Добавить _isZeroCalorieName guard в _parseText + IngredientV2.fromApiItem"),

    # ─ Незакрытые фидбеки пользователей ─
    ("4.1", "User feedback", "Грубые ошибки калорий USDA", "Open",
     "USDA mapping (бэк)",
     "Заменить таблицу USDA; добавить валидацию диапазонов"),
    ("8.1", "User feedback", "Долгое обновление стартового экрана", "Open",
     "dashboard_provider.dart",
     "Профилировать загрузку; добавить кэш + skeleton"),
    ("6.1", "User feedback", "Нельзя удалить блюдо", "Open",
     "meals API + journal UI",
     "DELETE /api/meals/<id> + swipe-to-delete в журнале"),
    ("6.3", "User feedback", "Нет голос/фото ввода в правках блюда", "Open",
     "edit_meal_screen.dart",
     "Переиспользовать AddMeal voice/photo flow внутри редактора"),
    ("5.4", "User feedback", "Не редактируется вес блюда", "Open",
     "edit_meal_screen.dart",
     "TextField для веса + пересчёт калорий"),
    ("5.5", "User feedback", "Correct меняет всё блюдо вместо одного ингредиента", "Open",
     "edit_meal_screen.dart",
     "Ограничить scope коррекции выбранным ингредиентом"),
    ("3.7", "User feedback", "Расхождение калорий между экранами", "Open",
     "stats provider + meal aggregation",
     "Аудит pipeline суммирования"),
    ("3.6", "User feedback", "Скачок графика веса", "Open",
     "weight_chart.dart + provider",
     "Сглаживание + проверить пробелы данных"),
    ("3.5", "User feedback", "Экстремальный целевой вес без предупреждения", "Open",
     "way_to_goal screen",
     "BMI-warning при submit"),
    ("5.6", "User feedback", "Нет редактирования БЖУ", "Open",
     "edit_meal_screen.dart",
     "4 TextField для protein/fat/carbs/calories"),
    ("5.3", "User feedback", "AI завышает калории без контекста", "Open",
     "ai_service_v2.py prompt",
     "Промпт: требовать визуальное подтверждение"),
    ("FAV", "User feedback", "Теги нелюбимых продуктов", "Open",
     "settings + onboarding",
     "Новый экран + persistent storage"),
    ("9.1", "User feedback", "«Помогите нам расти» в онбординге", "Open",
     "onboarding finale",
     "Экран запроса отзыва"),
    ("9.2", "User feedback", "Плашка отзыва в Settings", "Open",
     "settings_screen.dart",
     "CTA-баннер с переходом на review"),

    # ─ Новая большая задача ─
    ("NEW-OB", "New scope", "Переделать онбординг по FitnesBot", "Not started",
     "github.com/IzeBN/FitnesBot → onboarding/*",
     "Заимствовать экран «путь к цели»: выбор формы тела → процент жира → желаемый % → расчёт веса. + экраны «после достижения цели»"),

    # ─ Kayfit 2.0 redesign (radically minimal) ─
    ("KF2-FOUND-1", "Kayfit 2.0", "Дизайн-токены + KayfitRings widget", "Code-only",
     "lib/shared/theme/kayfit2_theme.dart, lib/shared/widgets/kayfit_rings.dart, lib/features/kayfit2/screens/kayfit2_preview_screen.dart",
     "Done в коде: 4-ring Apple Activity widget с CustomPainter, токены light+dark. Preview по /kayfit2/preview"),
    ("KF2-FOUND-2", "Kayfit 2.0", "Calendar strip + month grid widget", "Open",
     "новый: lib/shared/widgets/kayfit2_calendar_strip.dart",
     "Apple-style status rings (green=good, red=over) на week strip и month grid; раскрытие на тап"),
    ("KF2-FOUND-3", "Kayfit 2.0", "Tab bar Apple-style с центральной +", "Open",
     "новый: lib/shared/widgets/kayfit2_tab_bar.dart",
     "frosted glass + blur, blue accent #007AFF, центральная gradient +-кнопка 44pt, tap → открывает Chat"),
    ("KF2-FOUND-4", "Kayfit 2.0", "Meal photo placeholder + meal row", "Open",
     "новый: lib/shared/widgets/meal_row_v2.dart, meal_photo_placeholder.dart",
     "Striped greyscale tile 56x56 если photo, иначе time/source column. Полная meal row Apple-стиль"),
    ("KF2-JOURNAL", "Kayfit 2.0", "JournalV2Screen — собрать всё", "Open",
     "lib/features/journal/screens/journal_v2_screen.dart",
     "TopBar + CalendarStrip + KayfitRings + список MealRow + TabBar"),
    ("KF2-CHAT", "Kayfit 2.0", "ChatV2Screen с inline input + thinking bubble", "Open",
     "lib/features/chat/screens/chat_v2_screen.dart",
     "Сообщения + ai-thinking-bubble (parsing → USDA → FatSecret → compile) + inline input bar с camera/mic/barcode/send"),
    ("KF2-RECOG", "Kayfit 2.0", "Recognition capture/recognizing/preview screens", "Open",
     "lib/features/add_meal/screens/recognition_v2_screen.dart",
     "Apple-style: ✕ + label, capture circle 64pt, dashed→solid border, recognizing state"),
    ("KF2-PREVIEW-EDIT", "Kayfit 2.0", "Preview & edit с +/×/✏ + AI-correct", "Open",
     "lib/features/add_meal/widgets/preview_item_card.dart",
     "Per-item: pencil → P/F/C edit, X → delete, + → add. AI-correct natural language. kcal автопересчёт по 4/9/4"),
    ("KF2-CHAT-AI-PIPELINE", "Kayfit 2.0", "Реальный stream прогресс ИИ в чат", "Open",
     "Backend SSE/streaming + frontend listener",
     "Сейчас фронтовая симуляция — заменить на реальный SSE из ai_service_v2 (parsing → DB → compile)"),
    ("KF2-ROUTE", "Kayfit 2.0", "Roadmap routing — отказ от Dashboard", "Open",
     "lib/router.dart, ShellRoute",
     "Только 2 stack: /journal, /chat. / редиректит на /journal. /dashboard, /settings перевести в pop-over или тоже /journal"),
    ("KF2-FONTS", "Kayfit 2.0", "Подключить Geist + JetBrains Mono в pubspec", "Open",
     "pubspec.yaml + assets/fonts/",
     "Скачать шрифты под лицензией (Geist OFL, JBM OFL), положить в assets, прописать family"),
    ("KF2-MIGRATE", "Kayfit 2.0", "Cleanup старых экранов после переезда", "Open",
     "удалить dashboard_screen.dart, bottom_nav.dart (старый), add_meal_sheet.dart (modal grid)",
     "После того как KF2 проходит ручной QA — снести старые экраны и assets"),

    # ─ Plan Ready screen (Bug + redesign) ─
    ("PR-1", "Plan Ready", "Дубль экрана Your Plan Is Ready (онбординг + /way-to-goal)", "Done",
     "lib/features/auth/screens/email_auth_screen.dart",
     "Manual PASS (2.2): онбординг → логин → сразу AI consent / Dashboard, без повтора Plan Ready"),
    ("PR-2", "Plan Ready", "График наклон при current==target", "Done",
     "lib/features/way_to_goal/widgets/plan_result_view.dart",
     "Manual PASS (1.14): равные веса 70/70 → нет графика, есть баннер 'Поддержание веса'"),
    ("PR-3", "Plan Ready", "Frontend: баннер персонального AI-плана над hero", "Code-only",
     "lib/features/way_to_goal/widgets/plan_result_view.dart, lib/shared/models/calculation_result.dart",
     "Stub готов: читает calc.personalizedPlan, скрывает баннер если null. Покажется когда бэкенд начнёт возвращать поле"),
    ("PR-4-BE", "Plan Ready", "Backend: эндпоинт персонального плана (≤150 chars)", "Open",
     "backend: новый эндпоинт + AI prompt",
     "Сделать ai_call(diet, restrictions, goal) → 150-char personalized plan. Включить в /api/calculation/result как personalized_plan. Промпт: краткий, мотивирующий, с учётом ограничений"),

    # ─ Tech-debt ─
    ("TD-1", "Tech-debt", "3 failing tests journal_new_ui_test.dart", "Open",
     "test/journal_new_ui_test.dart",
     "Починить MealGroupCard / DaySummaryCard / NutrientColors — pre-existing"),
    ("TD-2", "Tech-debt", "TokenStorage deprecated shim", "Open",
     "lib/core/api/api_client.dart",
     "Удалить shim после миграции всех вызовов"),
    ("TD-3", "Tech-debt", "Localizations.localeOf вместо localeProvider", "Open",
     "chat_screen.dart, dashboard_screen.dart, recognition_result_sheet_v2.dart, add_meal_sheet.dart",
     "Мигрировать на ref.watch(localeProvider)"),
    ("TD-4", "Tech-debt", "Хардкод-строки без arb", "Open",
     "5+ экранов",
     "Полный аудит хардкода + извлечение в arb"),
    ("TD-5", "Tech-debt", "Fastlane/EAS release pipeline", "Open",
     "ios/ + android/",
     "Wire CI + signing automation"),

    # ─ Релизные требования ─
    ("REL-1", "Release", "iOS Keychain entitlements в release build", "Code-only",
     "ios/Runner/Runner.entitlements",
     "Manual: flutter build ipa --release → запуск на iPhone без crash"),
    ("REL-2", "Release", "Закоммитить hotfix-артефакты", "Open",
     "ios/Podfile.lock, ios/Runner/Runner.entitlements, specs/*.md",
     "git add + commit после прохождения ручной регрессии"),
    ("REL-3", "Release", "Merge hotfix-ветки в master", "Open",
     "claude/hotfix-p0-feedback-2026-05-02 → master",
     "После ручной регрессии: PR + review + merge"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Roadmap"

    headers = ["ID", "Категория", "Задача", "Статус", "Файлы / источник", "Что сделать"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    for it in ITEMS:
        ws.append(list(it))
        r = ws.max_row
        fill_status(ws.cell(row=r, column=4), it[3])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    # Column widths
    widths = [10, 18, 50, 14, 50, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(ITEMS)} items)")


if __name__ == "__main__":
    main()
